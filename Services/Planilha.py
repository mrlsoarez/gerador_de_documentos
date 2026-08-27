# leitura de planilha


#import win32com.client

#import pandas as pd
#import os

#from datetime import datetime
from openpyxl import load_workbook

#import time

# Funções referentes a planilha


def abrir_planilha(localizacao_planilha, sh):
    ANALISE = load_workbook(localizacao_planilha, data_only= True) 
    SHEET = ANALISE[sh]
    return SHEET 
    

"""

def ATUALIZAR_PLANILHA_COM_DADOS_EXTERNOS(localizacao_planilha, nome_sheet, dados):

    ANALISE_FISCAL = load_workbook(localizacao_planilha)
    
    def limpar_planilha(sheet):
        if sheet.max_row > 0:
            sheet.delete_rows(1, sheet.max_row)
    
    SHEET = ANALISE_FISCAL[nome_sheet]
    limpar_planilha(SHEET)

    HEAD = []
    for keys in dados[0]:
        HEAD.append(keys)
    SHEET.append(HEAD)

    for linha in dados: 
        SHEET.append([linha.get(coluna, "") for coluna in HEAD])
        
    ANALISE_FISCAL.save(localizacao_planilha)



    def atualizar_planilha(sheet_origem, sheet_destino):
        origem = sheet_origem.active
        destino = ANALISE_FISCAL[sheet_destino]
       
        for linha_idx, linha in enumerate(origem.iter_rows(values_only=True), start=1):
            for coluna_idx, valor in enumerate(linha, start=1):
                destino.cell(
                    row=linha_idx,
                    column=coluna_idx,
                    value=valor
                )
    
    try:
        converter_para_xlsx(rf"{pasta}\Portal Transparencia Despesas Gerais - Exercício 2026.xls")
        converter_para_xlsx(rf"{pasta}\Portal Transp. Despesas Liquidadas.xls")
    except: 
        pass

    ANALISE_FISCAL = load_workbook(localizacao_planilha)
    planilha_empenhos = load_workbook(rf"{pasta}\Portal Transparencia Despesas Gerais - Exercício 2026.xlsx")
    planilha_liquidacao = load_workbook(rf"{pasta}\Portal Transp. Despesas Liquidadas.xlsx")

    try:
        atualizar_planilha(planilha_empenhos, "Empenhos")
        atualizar_planilha(planilha_liquidacao, "Liquidacoes")
    except:
        print("Não foi possível completar a substituição dos dados na planilha.")
    else:
        ANALISE_FISCAL.save(localizacao_planilha)


     
def INICIAR_PLANILHA(localizacao_planilha, gerenciador_arquivos, param):

    def ouvir_planilha_incremental():
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        
        wb = excel.Workbooks.Open(localizacao_planilha)
        modified = datetime.fromtimestamp(os.path.getmtime(localizacao_planilha))
        
        print("■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■\nOUVINDO PLANILHA....\nAperte CTRL + C para encerrar a planilha\n■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■")
        
        try:
            while True:
                time.sleep(1)
                last_modified = os.path.getmtime(localizacao_planilha)
                last_modified = datetime.fromtimestamp(last_modified)
                if (last_modified > modified):
                    PROCESSAR_ARQUIVO(localizacao_planilha, gerenciador_arquivos, param)
                    modified = datetime.fromtimestamp(os.path.getmtime(localizacao_planilha))     
        except KeyboardInterrupt:
            print("Encerrando...\n■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■")
        finally:
            wb.Close(SaveChanges=True)  
            excel.Quit()               

    
    if ( (param["op"] == 1 and param["modo_total"]) or (param["op"] == 2) or param["op"] == 3): 
        print(param["op"], "a")
        return PROCESSAR_ARQUIVO(localizacao_planilha, gerenciador_arquivos, param)
    else: 
        ouvir_planilha_incremental()
   





    """